"""엑셀 출력 모듈."""
import calendar
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SHIFT_COLORS = {
    'D': '6FA8DC',  # 파랑
    'E': 'FFCE47',  # 노랑
    'N': 'A47BE0',  # 보라
    'O': 'E0E0E0',  # 회색
}
SHIFT_KR = {'D': '낮', 'E': '저녁', 'N': '밤', 'O': '휴'}

thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center')


def export(year: int, month: int, nurses: list, schedule: dict, path: str):
    days = calendar.monthrange(year, month)[1]
    weekdays = [calendar.weekday(year, month, d) for d in range(1, days + 1)]
    LBL = 3  # 라벨 열 수: 이름 / 사번 / 직책

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # ── 헤더 ──────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LBL)
    ws['A1'] = f"{year}년 {month}월 간호사 근무 스케줄"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = CENTER

    # 날짜 행
    ws.cell(row=2, column=1, value='이름')
    ws.cell(row=2, column=2, value='사번')
    ws.cell(row=2, column=3, value='직책')
    for d in range(1, days + 1):
        col = d + LBL
        cell = ws.cell(row=2, column=col, value=d)
        cell.alignment = CENTER
        cell.font = Font(bold=True)
        cell.border = BORDER
        # 주말 표시
        if weekdays[d - 1] == 5:  # 토
            cell.fill = PatternFill('solid', fgColor='9FCAF2')
        elif weekdays[d - 1] == 6:  # 일
            cell.fill = PatternFill('solid', fgColor='FFAFAF')

    # 합계 열 헤더
    for i, lbl in enumerate(['낮합', '저녁합', '밤합', '휴합', '근무합']):
        col = days + LBL + 1 + i
        cell = ws.cell(row=2, column=col, value=lbl)
        cell.alignment = CENTER
        cell.font = Font(bold=True)
        cell.border = BORDER

    # 요일 행
    wd_kr = ['월', '화', '수', '목', '금', '토', '일']
    ws.cell(row=3, column=1, value='요일')
    for d in range(1, days + 1):
        cell = ws.cell(row=3, column=d + LBL, value=wd_kr[weekdays[d - 1]])
        cell.alignment = CENTER
        cell.border = BORDER
        if weekdays[d - 1] == 5:
            cell.fill = PatternFill('solid', fgColor='9FCAF2')
        elif weekdays[d - 1] == 6:
            cell.fill = PatternFill('solid', fgColor='FFAFAF')

    # ── 데이터 행 ──────────────────────────────────────────────────────────────
    day_totals = {d: {'D': 0, 'E': 0, 'N': 0} for d in range(1, days + 1)}

    for row_idx, nurse in enumerate(nurses):
        row = row_idx + 4
        ws.cell(row=row, column=1, value=nurse.name).border = BORDER
        ws.cell(row=row, column=2, value=getattr(nurse, 'emp_no', '') or '').border = BORDER
        ws.cell(row=row, column=3, value=nurse.position).border = BORDER

        counts = {'D': 0, 'E': 0, 'N': 0, 'O': 0}
        nurse_sched = schedule.get(nurse.id, {})
        for d in range(1, days + 1):
            shift = nurse_sched.get(d, 'O')
            counts[shift] += 1
            if shift in day_totals[d]:
                day_totals[d][shift] += 1

            cell = ws.cell(row=row, column=d + LBL, value=SHIFT_KR[shift])
            cell.alignment = CENTER
            cell.border = BORDER
            cell.fill = PatternFill('solid', fgColor=SHIFT_COLORS[shift])

        # 합계
        for i, key in enumerate(['D', 'E', 'N', 'O']):
            ws.cell(row=row, column=days + LBL + 1 + i, value=counts[key]).alignment = CENTER
        ws.cell(row=row, column=days + LBL + 5,
                value=counts['D'] + counts['E'] + counts['N']).alignment = CENTER

    # ── 일별 합계 행 ──────────────────────────────────────────────────────────
    sum_row = len(nurses) + 4
    ws.cell(sum_row, 1, '낮합계').font = Font(bold=True)
    ws.cell(sum_row + 1, 1, '저녁합계').font = Font(bold=True)
    ws.cell(sum_row + 2, 1, '밤합계').font = Font(bold=True)
    for d in range(1, days + 1):
        ws.cell(sum_row,     d + LBL, day_totals[d]['D']).alignment = CENTER
        ws.cell(sum_row + 1, d + LBL, day_totals[d]['E']).alignment = CENTER
        ws.cell(sum_row + 2, d + LBL, day_totals[d]['N']).alignment = CENTER

    # ── 열 너비 조정 ──────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12   # 사번
    ws.column_dimensions['C'].width = 7
    for d in range(1, days + 1):
        ws.column_dimensions[get_column_letter(d + LBL)].width = 4
    for i in range(5):
        ws.column_dimensions[get_column_letter(days + LBL + 1 + i)].width = 7

    ws.freeze_panes = 'D4'
    wb.save(path)
    return path
